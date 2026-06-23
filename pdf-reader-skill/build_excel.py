# -*- coding: utf-8 -*-
"""Baut die Schulden-Uebersicht (Excel, 6 Blaetter) aus result.json + manifest.json."""
import json, pathlib, re, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEUTE = datetime.date(2026, 6, 16)
BASE = pathlib.Path(r"C:\Users\KON-01\.claude\skills\pdf-reader-skill")
res = json.loads((BASE / "result.json").read_text(encoding="utf-8"))
man = {m["id"]: m["name"] for m in json.loads(pathlib.Path(r"C:\Users\KON-01\pdf_txt\manifest.json").read_text(encoding="utf-8"))}

# ---------- Cluster-Zuordnung ----------
def cluster(g, pdf):
    s = (g + " " + pdf).lower()
    rules = [
        ("WEG Karl-Knoedl / Raumgut (Hausgeld)", ["knoedl","knodl","knödl","raumgut","eigentuemergemein","eigentümergemein","weg karl","bws"]),
        ("RheinEnergie", ["rheinenergie"]),
        ("Riverty (Inkasso)", ["riverty"]),
        ("IONOS", ["ionos","1&1","1 1 telecom"]),
        ("Rundfunkbeitrag / WDR", ["rundfunk","westdeutscher rundfunk","wdr","beitragsservice"]),
        ("IHK Duesseldorf", ["handelskammer","ihk"]),
        ("Finanzamt Hilden", ["finanzamt"]),
        ("BFS Health Finance", ["bfs health","bfs"]),
        ("Sparkasse", ["sparkasse"]),
        ("Nuernberger Versicherung", ["nurnberger","nürnberger","nurenberg","nürnberg"]),
        ("VHV Versicherung", ["vhv"]),
        ("VW/Audi Leasing", ["volkswagen leasing","audi leasing","vw leasing","leasing"]),
        ("VW/Audi Versicherung (Lowell)", ["lowell","versicherungs ag","autoversicherung"]),
        ("HUK-Coburg", ["huk"]),
        ("HanseMerkur", ["hansemerkur","hanse merkur","merkur"]),
        ("ERGO", ["ergo"]),
        ("Zahnaerztliche Behandlung", ["zahn"]),
        ("Stadtwerke Langenfeld", ["stadtwerke langenfeld"]),
        ("Grundsteuer Duesseldorf", ["grundsteuer","stadtkasse","landeshauptstadt"]),
        ("Stadt Langenfeld", ["stadt langenfeld"]),
        ("American Express", ["american express","amex"]),
        ("Zentrale Justiz / Zahlstelle", ["justiz","zahlstelle"]),
        ("Fitness Gym Erkrath", ["fitness","gym"]),
        ("Steuerberater Meiners & Eulers", ["steuerberater","stbvs","meiners","eulers"]),
        ("Bueromiete proImmobilie", ["proimmobilie","radszun","pro immobilien"]),
        ("Rheinbahn", ["rheinbahn"]),
        ("PayPal / KSP", ["paypal","ksp"]),
        ("Anwalt Familienrecht", ["lauppe","hasskamp","familienrecht"]),
        ("Anwalt Arbeitsgericht", ["arbeitsgericht"]),
        ("Alte Leipziger", ["leipziger"]),
        ("Private Zusatzversicherung (ARAG/AR)", ["zusatzversicherung","arag","z100"]),
        ("Domain Kaufbuddy", ["kaufbuddy","domain"]),
        ("Audi Reifeneinlagerung", ["reifen"]),
        ("Automobil Zentrum Leverkusen", ["automobil zentrum","leverkusen"]),
        ("Techniker Krankenkasse", ["techniker krankenkasse","tk "]),
        ("Strassenverkehrsamt", ["strassenverkehr","straßenverkehr","ordnungswidrig","festsetzung"]),
        ("Bundesagentur fuer Arbeit", ["bundesagentur","arbeitsagentur"]),
    ]
    for name, keys in rules:
        if any(k in s for k in keys):
            return name
    # Fallback: erste 3 Woerter des Glaeubigers
    return " ".join(re.sub(r"[^A-Za-zÄÖÜäöüß ]"," ", g).split()[:3]) or "Sonstiges"

# ---------- Kategorie ----------
def kategorie(f, pdf):
    p = pdf.lower(); st = (f.get("status") or "").lower()
    if p.startswith("bezahlt"):
        return "Vermutlich bezahlt (Ablage)"
    if "guthaben" in p or "gutschrift" in st:
        return "Guthaben/Gutschrift"
    if "stoniert" in p or "storniert" in p:
        return "Storniert"
    if f.get("betrag_eur") is None:
        return "Ohne bezifferten Betrag"
    if st in ("zahlung","information"):
        return "Information/Nachweis"
    return "Offen"

def to_date(s):
    if not s: return None
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", s)
    if not m: return None
    try: return datetime.date(int(m[3]), int(m[2]), int(m[1]))
    except: return None

# ---------- Zeilen aufbauen ----------
rows = []
for d in res["dokumente"]:
    pdf = man.get(d["id"], d["id"])
    for f in d.get("forderungen", []):
        rows.append({
            "id": d["id"], "pdf": pdf, "seite": f.get("seite"),
            "glaeubiger": f.get("glaeubiger",""), "betrag": f.get("betrag_eur"),
            "haupt": f.get("hauptforderung"), "kosten": f.get("kosten"), "zinsen": f.get("zinsen"),
            "datum": f.get("datum_dokument"), "faellig": f.get("faellig"), "frist": f.get("frist_bis"),
            "status": f.get("status",""), "fp": f.get("firma_oder_privat",""),
            "typ": f.get("glaeubiger_typ",""), "kassenzeichen": f.get("kassenzeichen"),
            "gegenstand": f.get("gegenstand",""), "flags": f.get("flags"),
            "cluster": cluster(f.get("glaeubiger",""), pdf),
        })
        rows[-1]["kategorie"] = kategorie(f, pdf)

for i, r in enumerate(rows, 1):
    r["nr"] = i

# ---------- Duplikat-/Eskalations-Bereinigung ----------
# Explizite Reduktionen (id-basiert) fuer eindeutige Doppelzaehlungen:
DUP_REMOVE = {
    # nr-unabhaengig: (cluster, betrag) Treffer werden als Duplikat markiert
}
# WEG: alle "Offen"-Posten werden NICHT summiert; stattdessen ein massgeblicher Betrag.
CLUSTER_MASSGEBLICH = {
    "WEG Karl-Knoedl / Raumgut (Hausgeld)": (8191.55, "Hoechste aktuelle Gesamtforderung (BWS, 25.03.2026, inkl. Kosten/Zinsen) im Zwangsversteigerungsverfahren 84 K 1/26. Frueheres Stadium 5.512,19 EUR. Exakten Tagesstand bei Gericht/raumgut erfragen. Spanne 5.512-8.192 EUR."),
}
# Eindeutige Voll-Duplikate (gleicher Betrag, gleiche Schuld in zwei Dateien) -> eine Zeile als Duplikat:
def is_duplicate(r):
    b = r["betrag"]
    # Rundfunk-Vollstreckung 181,24 doppelt (052 dupliziert 076)
    if r["id"] == "052" and b == 181.24:
        return "Duplikat zu Dok 076 (gleiche Rundfunk-Vollstreckung 181,24)"
    # American Express 1030,01 doppelt (056 via Inkasso = 066)
    if r["id"] == "056" and b == 1030.01:
        return "Duplikat zu Dok 066 (gleiche Amex-Forderung 1.030,01)"
    return None

for r in rows:
    dup = is_duplicate(r)
    r["duplikat"] = dup
    # In WEG-Cluster zaehlt kein Einzelposten zur bereinigten Summe (separat ersetzt)
    massgeblich = r["cluster"] in CLUSTER_MASSGEBLICH
    r["zaehlt_offen"] = (r["kategorie"] == "Offen" and not dup and not massgeblich)

# ---------- Summen ----------
def euros(rs): return round(sum(r["betrag"] for r in rs if isinstance(r["betrag"], (int,float))), 2)
offen_rows = [r for r in rows if r["kategorie"] == "Offen"]
brutto_offen = euros(offen_rows)
bereinigt_rows = [r for r in rows if r["zaehlt_offen"]]
bereinigt_offen = euros(bereinigt_rows) + sum(v[0] for v in CLUSTER_MASSGEBLICH.values())
firma_offen = euros([r for r in bereinigt_rows if r["fp"] == "Firma"]) + sum(v[0] for c,v in CLUSTER_MASSGEBLICH.items())  # WEG ist privat -> unten korrigiert
# WEG ist Privat: massgeblich zu Privat zaehlen
firma_offen = euros([r for r in bereinigt_rows if r["fp"] == "Firma"])
privat_offen = euros([r for r in bereinigt_rows if r["fp"] == "Privat"]) + sum(v[0] for v in CLUSTER_MASSGEBLICH.values())
unklar_offen = euros([r for r in bereinigt_rows if r["fp"] not in ("Firma","Privat")])
bezahlt_sum = euros([r for r in rows if r["kategorie"].startswith("Vermutlich bezahlt")])
gutschrift_sum = euros([r for r in rows if r["kategorie"] == "Guthaben/Gutschrift"])

# ---------- Excel ----------
wb = Workbook()
HEAD = Font(bold=True, color="FFFFFF", size=10)
HFILL = PatternFill("solid", fgColor="1F4E78")
TITLE = Font(bold=True, size=14, color="1F4E78")
BOLD = Font(bold=True)
RED = PatternFill("solid", fgColor="F8CBAD")
YELLOW = PatternFill("solid", fgColor="FFE699")
GREEN = PatternFill("solid", fgColor="C6E0B4")
GREY = PatternFill("solid", fgColor="D9D9D9")
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
EUR = '#,##0.00 "EUR"'

def style_header(ws, ncol, rowidx=1):
    for c in range(1, ncol+1):
        cell = ws.cell(row=rowidx, column=c)
        cell.font = HEAD; cell.fill = HFILL; cell.alignment = WRAP; cell.border = BORD
    ws.freeze_panes = ws.cell(row=rowidx+1, column=1)

# ---- Blatt 1: Zusammenfassung ----
ws = wb.active; ws.title = "Zusammenfassung"
ws["A1"] = "Schuldenuebersicht Jerome Didelet"; ws["A1"].font = TITLE
ws["A2"] = f"Stand der Auswertung: {HEUTE.strftime('%d.%m.%Y')}  |  91 PDF-Dokumente, {len(rows)} erfasste Posten"
ws["A2"].font = Font(italic=True, color="808080")
data = [
    ("", ""),
    ("KENNZAHLEN (Schaetzung - vom Berater zu pruefen)", ""),
    ("Offene Schulden - BRUTTO (alle offenen Einzelposten, mit moeglichen Doppelzaehlungen)", brutto_offen),
    ("Offene Schulden - BEREINIGT (Eskalationen/Duplikate zusammengefasst)", bereinigt_offen),
    ("   davon FIRMA (Konstrukt GmbH i.L.)", firma_offen),
    ("   davon PRIVAT (Jerome Didelet)", privat_offen),
    ("   davon UNKLAR (Zuordnung offen)", unklar_offen),
    ("", ""),
    ("Vermutlich bereits BEZAHLT (Ablage 'Bezahlt...', zu bestaetigen)", bezahlt_sum),
    ("Guthaben/Gutschriften (zu Gunsten Didelet)", gutschrift_sum),
    ("", ""),
    ("TOP-PRIORITAETEN (groesste / dringendste offene Posten)", ""),
]
ridx = 4
for label, val in data:
    ws.cell(row=ridx, column=1, value=label)
    if isinstance(val, (int,float)):
        c = ws.cell(row=ridx, column=2, value=val); c.number_format = EUR; c.font = BOLD
    if label.startswith("Offene Schulden - BEREINIGT"): ws.cell(row=ridx,column=1).font = BOLD; ws.cell(row=ridx,column=2).fill = RED
    if label.startswith("Offene Schulden - BRUTTO"): ws.cell(row=ridx,column=1).font = BOLD
    if "KENNZAHLEN" in label or "TOP-PRIO" in label: ws.cell(row=ridx,column=1).font = Font(bold=True, color="1F4E78", size=11)
    ridx += 1
# Top-Positionen
tops = sorted([r for r in rows if r["kategorie"]=="Offen" and isinstance(r["betrag"],(int,float))], key=lambda r:-r["betrag"])[:10]
ws.cell(row=ridx, column=1, value="Gläubiger / Cluster").font = BOLD
ws.cell(row=ridx, column=2, value="Betrag").font = BOLD
ws.cell(row=ridx, column=3, value="Status / Frist").font = BOLD
ridx += 1
for r in tops:
    ws.cell(row=ridx, column=1, value=f"[{r['id']}] {r['cluster']}")
    c = ws.cell(row=ridx, column=2, value=r["betrag"]); c.number_format = EUR
    fr = to_date(r["frist"]); fr_s = f"Frist {r['frist']}" if r["frist"] else r["status"]
    if fr and fr < HEUTE: fr_s += " (ueberschritten!)"; c.fill = RED
    ws.cell(row=ridx, column=3, value=f"{r['status']} | {fr_s}")
    ridx += 1
ws.column_dimensions["A"].width = 62; ws.column_dimensions["B"].width = 16; ws.column_dimensions["C"].width = 34

# ---- Blatt 2: Übersicht (alle Posten) ----
ws = wb.create_sheet("Übersicht")
cols = ["Nr","Dok-ID","PDF-Quelle","Seite","Gläubiger","Cluster","Betrag (EUR)","Hauptf.","Kosten","Zinsen",
        "Dok-Datum","Fällig","Frist","Tage b. Frist","Status","Kategorie","Firma/Privat","Typ","Kassenzeichen","Gegenstand","Flags/Hinweise"]
ws.append(cols)
for r in rows:
    fr = to_date(r["frist"]); tage = (fr - HEUTE).days if fr else None
    ws.append([r["nr"], r["id"], r["pdf"], r["seite"], r["glaeubiger"], r["cluster"], r["betrag"],
               r["haupt"], r["kosten"], r["zinsen"], r["datum"], r["faellig"], r["frist"], tage,
               r["status"], r["kategorie"], r["fp"], r["typ"], r["kassenzeichen"], r["gegenstand"],
               (r["duplikat"]+" | " if r["duplikat"] else "") + (r["flags"] or "")])
    rr = ws.max_row
    for col in (7,8,9,10): ws.cell(row=rr,column=col).number_format = EUR
    kat = r["kategorie"]
    fill = None
    if r["duplikat"]: fill = GREY
    elif kat == "Offen":
        if tage is not None and tage < 0: fill = RED
    elif kat.startswith("Vermutlich bezahlt") or kat=="Guthaben/Gutschrift": fill = GREEN
    elif kat in ("Information/Nachweis","Storniert","Ohne bezifferten Betrag"): fill = GREY
    if fill:
        for c in range(1, len(cols)+1): ws.cell(row=rr, column=c).fill = fill
style_header(ws, len(cols))
widths = [4,6,34,7,30,28,13,11,11,11,11,11,11,9,16,20,11,16,24,46,40]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

# ---- Blatt 3 + 4: Firma / Privat (nur Offen + bereinigt) ----
def sheet_fp(name, fp):
    ws = wb.create_sheet(name)
    cols = ["Nr","Dok-ID","Gläubiger / Cluster","Betrag (EUR)","Dok-Datum","Frist","Status","Kassenzeichen","Gegenstand"]
    ws.append(cols)
    sub = [r for r in rows if r["fp"]==fp and r["zaehlt_offen"]]
    sub.sort(key=lambda r: -(r["betrag"] or 0))
    for r in sub:
        ws.append([r["nr"], r["id"], f"{r['cluster']} — {r['glaeubiger'][:40]}", r["betrag"], r["datum"], r["frist"], r["status"], r["kassenzeichen"], r["gegenstand"][:80]])
        ws.cell(row=ws.max_row, column=4).number_format = EUR
    # WEG massgeblich bei Privat ergaenzen
    extra = 0
    if fp == "Privat":
        for c,(v,note) in CLUSTER_MASSGEBLICH.items():
            ws.append(["—","div.", c + " (massgeblicher Gesamtbetrag)", v, "", "", "ESKALATION", "84 K 1/26", note[:80]])
            ws.cell(row=ws.max_row, column=4).number_format = EUR
            ws.cell(row=ws.max_row, column=4).fill = RED; extra += v
    total = euros(sub) + extra
    ws.append([]); ws.append(["","","SUMME "+name+" (offen, bereinigt)", total]);
    ws.cell(row=ws.max_row, column=3).font = BOLD
    c = ws.cell(row=ws.max_row, column=4); c.number_format = EUR; c.font = BOLD; c.fill = YELLOW
    style_header(ws, len(cols))
    for i,w in enumerate([4,6,46,13,11,11,18,24,60],1): ws.column_dimensions[get_column_letter(i)].width = w
sheet_fp("Firma", "Firma")
sheet_fp("Privat", "Privat")

# ---- Blatt 5: Cluster & Duplikate ----
ws = wb.create_sheet("Cluster & Duplikate")
ws.append(["Cluster","Posten","Summe offene Einzelposten (EUR)","Bewertung / massgeblicher Betrag","Hinweis"])
from collections import defaultdict
cl = defaultdict(list)
for r in rows: cl[r["cluster"]].append(r)
for name in sorted(cl, key=lambda n:-len(cl[n])):
    g = cl[name]
    offen = [r for r in g if r["kategorie"]=="Offen"]
    s = euros(offen)
    if name in CLUSTER_MASSGEBLICH:
        v,note = CLUSTER_MASSGEBLICH[name]
        bew = f"ESKALATION einer Schuld -> {v:.2f} EUR massgeblich"; hint = note
    elif len(g) > 1:
        bew = "Mehrere Posten - pruefen ob separat oder Eskalation";
        hint = "; ".join(f"[{r['id']}] {r['betrag']} {r['status']}" for r in g[:6] if r["betrag"])
    else:
        bew = "Einzelposten"; hint = (g[0]["duplikat"] or "")
    ws.append([name, len(g), s, bew, hint])
    if name in CLUSTER_MASSGEBLICH: ws.cell(row=ws.max_row, column=4).fill = RED
    ws.cell(row=ws.max_row, column=3).number_format = EUR
style_header(ws, 5)
for i,w in enumerate([38,7,22,42,80],1): ws.column_dimensions[get_column_letter(i)].width = w

# ---- Blatt 6: Chronologie & Fristen ----
ws = wb.create_sheet("Chronologie & Fristen")
cols = ["Tage b. Frist","Frist","Dok-ID","Gläubiger / Cluster","Betrag (EUR)","Status","Firma/Privat","Prioritaet"]
ws.append(cols)
mit_frist = [r for r in rows if r["kategorie"]=="Offen" and to_date(r["frist"])]
mit_frist.sort(key=lambda r: to_date(r["frist"]))
for r in mit_frist:
    fr = to_date(r["frist"]); tage = (fr - HEUTE).days
    if tage < 0: prio = "🔴 ÜBERSCHRITTEN"
    elif tage <= 7: prio = "🔴 KRITISCH (<7T)"
    elif tage <= 30: prio = "🟠 DRINGEND (<30T)"
    elif tage <= 90: prio = "🟡 WARNUNG"
    else: prio = "⚪ später"
    ws.append([tage, r["frist"], r["id"], f"{r['cluster']}", r["betrag"], r["status"], r["fp"], prio])
    ws.cell(row=ws.max_row, column=5).number_format = EUR
    if tage < 0:
        for c in range(1,len(cols)+1): ws.cell(row=ws.max_row,column=c).fill = RED
    elif tage <= 30:
        for c in range(1,len(cols)+1): ws.cell(row=ws.max_row,column=c).fill = YELLOW
style_header(ws, len(cols))
for i,w in enumerate([13,11,6,40,13,18,11,20],1): ws.column_dimensions[get_column_letter(i)].width = w

out = pathlib.Path(r"C:\Users\KON-01\Desktop\scan\Übersicht Unterlagen\001 Fertig\Schuldenuebersicht_Didelet.xlsx")
wb.save(out)
print("GESPEICHERT:", out)
print(f"Posten gesamt: {len(rows)}")
print(f"BRUTTO offen: {brutto_offen:,.2f} EUR")
print(f"BEREINIGT offen: {bereinigt_offen:,.2f} EUR  (Firma {firma_offen:,.2f} | Privat {privat_offen:,.2f} | Unklar {unklar_offen:,.2f})")
print(f"Vermutlich bezahlt: {bezahlt_sum:,.2f} EUR | Gutschriften: {gutschrift_sum:,.2f} EUR")
